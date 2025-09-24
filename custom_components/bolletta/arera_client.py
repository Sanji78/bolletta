"""ARERA Excel file parser for retrieving tariff parameters."""

import io
import logging
from datetime import date, datetime
from typing import Set, Dict, Optional, Any
import calendar
import re


import openpyxl
from aiohttp import ClientSession
from homeassistant.core import HomeAssistant
from homeassistant.helpers.aiohttp_client import async_get_clientsession

from .const import (
    RESIDENTIAL,
    NOT_RESIDENTIAL,
    HOUSE_TYPE_LABELS,
    CONF_ENERGY_SC1,
    CONF_FIX_QUOTA_TRANSPORT,
    CONF_QUOTA_POWER,
    CONF_ASOS_SC1,
    CONF_ARIM_SC1
)

_LOGGER = logging.getLogger(__name__)

# ARERA Excel URL template (updates quarterly)
ARERA_BASE_URL = "https://www.arera.it/fileadmin/area_operatori/prezzi_e_tariffe/"
ARERA_FILENAME_TEMPLATE = "E{year}_stg_domesticiNonVulnerabili.xlsx"


class AreraClient:
    """Client for downloading and parsing ARERA tariff data."""

    def __init__(self, hass: HomeAssistant):
        """Initialize the ARERA client."""
        self.hass = hass
        self.session: ClientSession = async_get_clientsession(hass)
        self._cached_data: Dict[str, Any] = {}
        self._cache_date: Optional[date] = None


    async def get_current_tariffs(self, house_type) -> Dict[str, Dict[str, float]]:
        """
        Get tariff parameters for current month and previous month from a single ARERA file.
        Returns a dict with up to two keys: { "mp": {...}, "mpp": {...} }
        """
        current_date = datetime.now().date()
        current_year = current_date.year
        current_month = current_date.month - 1
        if current_month == 0:
            current_month = 12
            current_year = current_year - 1
            
        # Calcolo mese precedente e anno corrispondente
        if current_month == 1:
            prev_month = 12
            prev_year = current_year - 1
        else:
            prev_month = current_month - 1
            prev_year = current_year

        # Target months e anni da parsare
        target_months = {(prev_year, prev_month), (current_year, current_month)}

        _LOGGER.info("ARERA: scarico il file excel, ricerco i mesi: %s", target_months)

        # Scarico e parso il file **una sola volta**
        filename = ARERA_FILENAME_TEMPLATE.format(year=current_year)
        url = f"{ARERA_BASE_URL}{filename}"
        _LOGGER.info("Scarico i dati ARERA da: %s", url)

        async with self.session.get(url) as response:
            if response.status != 200:
                raise RuntimeError(f"HTTP {response.status} nello scarico dei dati ARERA")
            content = await response.read()

        # Parsing sincronamente nel thread executor
        await self.hass.async_add_executor_job(
            self._parse_excel_data, content, target_months, house_type
        )

        # Costruisco il risultato con le chiavi 'mp' (mese-1) e 'mpp' (mese-2)
        mp_key = f"{current_year}_{current_month:02d}"
        mpp_key = f"{prev_year}_{prev_month:02d}"

        mp = self._cached_data.get(mp_key, {})
        mpp = self._cached_data.get(mpp_key, {})

        if not mp:
            _LOGGER.warning("Dati per 'mp' (%s) non trovati nel file ARERA.", mp_key)
        if not mpp:
            _LOGGER.warning("Dati per 'mpp' (%s) non trovati nel file ARERA.", mpp_key)

        # Ritorno finale con esattamente le chiavi richieste
        return {"mp": mp, "mpp": mpp}

    def _parse_excel_data(self, content: bytes, target_months: set[tuple[int,int]], house_type) -> None:
        """Parse only the sheets corresponding to target_months (year, month tuples)."""
        workbook = openpyxl.load_workbook(io.BytesIO(content), data_only=True)

        month_map = {
            "gen": 1, "gennaio": 1,
            "feb": 2, "febbraio": 2,
            "mar": 3, "marzo": 3,
            "apr": 4, "aprile": 4,
            "mag": 5, "maggio": 5,
            "giu": 6, "giugno": 6,
            "lug": 7, "luglio": 7,
            "ago": 8, "agosto": 8,
            "set": 9, "settembre": 9,
            "ott": 10, "ottobre": 10,
            "nov": 11, "novembre": 11,
            "dic": 12, "dicembre": 12,
        }

        for sheet_name in workbook.sheetnames:
            name_lower = sheet_name.lower()

            # Estrazione anno dal nome del foglio
            year_in_name = None
            for token in name_lower.split():
                if token.isdigit() and len(token) == 4:
                    year_in_name = int(token)
                    break

            # Estrazione mese dal nome del foglio
            found_month = None
            for prefix, num in month_map.items():
                if prefix in name_lower:
                    found_month = num
                    break

            if found_month is None or year_in_name is None:
                continue

            if (year_in_name, found_month) not in target_months:
                continue  # skip fogli non richiesti

            # estrazione parametri
            sheet = workbook[sheet_name]
            _LOGGER.debug("Accedo al foglio '%s'", sheet_name)
            month_data = self._extract_tariff_parameters(sheet, house_type)
            key = f"{year_in_name}_{found_month:02d}"
            self._cached_data[key] = month_data
            _LOGGER.debug("Ho trovato nel foglio '%s' -> %s = %s", sheet_name, key, month_data)


    def _extract_tariff_parameters(self, sheet, house_type) -> Dict[str, float]:
        parameters: Dict[str, float] = {}
        _LOGGER.debug("Utilizzo i dati per: '%s'", HOUSE_TYPE_LABELS.get(house_type))

        def _parse_numeric(val):
            import re
            if val is None:
                return None
            if isinstance(val, (int, float)):
                return float(val)
            s = str(val).strip()
            if s == '':
                return None
            try:
                if s.endswith('%'):
                    return float(s[:-1].replace(',', '.')) / 100.0
                s = s.replace(' ', '')
                s = s.replace(',', '.')
                s_clean = re.sub(r'[^\d\.\-eE]', '', s)
                if s_clean == '' or s_clean in ('-', '.', '+'):
                    return None
                return float(s_clean)
            except Exception:
                return None

        target_label = (HOUSE_TYPE_LABELS.get(house_type) or '').strip().lower()
        if not target_label:
            _LOGGER.warning("HOUSE_TYPE_LABELS.get(house_type) è vuoto o non definito.")
            return parameters

        found_label_row = None
        for r in range(1, sheet.max_row + 1):
            cell_val = sheet.cell(row=r, column=2).value
            if cell_val is None:
                continue
            if target_label in str(cell_val).strip().lower():
                found_label_row = r
                _LOGGER.debug("Trovato label '%s' in colonna 2 alla riga %d (valore cella: %s)",
                              target_label, r, repr(cell_val))
                break

        if not found_label_row:
            _LOGGER.warning("Label '%s' non trovata nella colonna 2 del foglio.", target_label)
            return parameters

        header_row = found_label_row + 1
        asos_arim_value_row = header_row + 3
        servizi_col = None

        for c in range(1, sheet.max_column + 1):
            header_cell = sheet.cell(row=header_row, column=c).value
            if header_cell is None:
                continue
            header_text = ' '.join(str(header_cell).split()).lower()

            # ASOS / ARIM (two rows below header_row)
            if 'asos' in header_text and CONF_ASOS_SC1 not in parameters:
                v = _parse_numeric(sheet.cell(row=asos_arim_value_row, column=c).value)
                if v is not None:
                    parameters[CONF_ASOS_SC1] = v
                    _LOGGER.debug("ASOS trovato in r%d c%d -> %s", asos_arim_value_row, c, v)
            if 'arim' in header_text and CONF_ARIM_SC1 not in parameters:
                v = _parse_numeric(sheet.cell(row=asos_arim_value_row, column=c).value)
                if v is not None:
                    parameters[CONF_ARIM_SC1] = v
                    _LOGGER.debug("ARIM trovato in r%d c%d -> %s", asos_arim_value_row, c, v)

            # detect "servizi di rete" even if it contains newlines or extra spaces
            if 'servizi' in header_text and 'rete' in header_text:
                servizi_col = c
                _LOGGER.debug("Intestazione 'Servizi di rete' trovata in c%d (testo intestazione: %s)", c, header_text)

        if servizi_col is not None:
            energy_row = header_row + 3
            fixq_row = header_row + 4
            quota_row = header_row + 5

            raw_energy = sheet.cell(row=energy_row, column=servizi_col).value
            raw_fixq = sheet.cell(row=fixq_row, column=servizi_col).value
            raw_quota = sheet.cell(row=quota_row, column=servizi_col).value

            v_energy = _parse_numeric(raw_energy)
            v_fixq = _parse_numeric(raw_fixq)
            v_quota = _parse_numeric(raw_quota)

            if v_energy is not None:
                parameters[CONF_ENERGY_SC1] = v_energy
                _LOGGER.debug("%s = %s (from r%d c%d)", CONF_ENERGY_SC1, v_energy, energy_row, servizi_col)
            else:
                _LOGGER.debug("%s non numerico in r%d c%d: %s", CONF_ENERGY_SC1, energy_row, servizi_col, repr(raw_energy))

            if v_fixq is not None:
                parameters[CONF_FIX_QUOTA_TRANSPORT] = v_fixq / 12.0
                _LOGGER.debug("%s = %s (raw %s from r%d c%d /12)", CONF_FIX_QUOTA_TRANSPORT, parameters[CONF_FIX_QUOTA_TRANSPORT], raw_fixq, fixq_row, servizi_col)
            else:
                _LOGGER.debug("%s non numerico in r%d c%d: %s", CONF_FIX_QUOTA_TRANSPORT, fixq_row, servizi_col, repr(raw_fixq))

            if v_quota is not None:
                parameters[CONF_QUOTA_POWER] = v_quota / 12.0
                _LOGGER.debug("%s = %s (raw %s from r%d c%d /12)", CONF_QUOTA_POWER, parameters[CONF_QUOTA_POWER], raw_quota, quota_row, servizi_col)
            else:
                _LOGGER.debug("%s non numerico in r%d c%d: %s", CONF_QUOTA_POWER, quota_row, servizi_col, repr(raw_quota))
        else:
            _LOGGER.debug("Colonna 'Servizi di rete' non trovata nella riga di intestazione %d", header_row)

        _LOGGER.debug("Parametri finali: %s", parameters)
        return parameters

    async def get_tariff_with_fallback(self, house_type) -> Dict[str, float]:
        """Get ARERA tariffs with fallback to defaults."""
        try:
            tariffs = await self.get_current_tariffs(house_type)
            if not tariffs:
                _LOGGER.warning("Non ci sono dati ARERA disponibili, uso quelli di default")
                return None
            return tariffs
            
        except Exception as e:
            _LOGGER.error("Non riesco ad ottenere i valori ARERA, uso quelli di default: %s", e)
            return None